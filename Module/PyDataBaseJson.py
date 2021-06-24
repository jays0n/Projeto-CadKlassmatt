import json
import openpyxl as xl
from PyDefines import *
import sys

def __concateDict__(dictA,dictB):
    dictResult=dictA
    dictResult.update(dictB)
    return dictResult

def checkKeys(key,dictA):
    if key in dictA.keys():
        return TRUE
    return FALSE


class manageJsonBd:
    #ATENÇÃO - Todos os metodos Update podem ocasionar sobrescrição de dados
    #Pay attention: All Update methods might cause data subscription.
    def __init__(self,filePath):
        self.pathFile=filePath
        self.createDataBase()

    def setMainPath(self,path):
        self.pathFile=path

    def createDataBase(self):
        try:
            stdContent=DCTNULL
            with open(self.pathFile,'x') as file:            
                json.dump(stdContent,file)
        except:
            pass

    def deleteJson(self):
        if os.path.exists(self.pathFile)==True:
            try:
                os.remove(self.pathFile)
                return TRUE,ERR_NO_ERRO,STRNULL
            except Exception as msgErr:
                return FALSE,ERR_FILE_UNKNOWN_ERROR,msgErr
        return TRUE,ERR_FILE_NOT_EXIST,STRNULL

    def updateJson(self,newBase):
        try:
            with open(self.pathFile,'w') as file:
                json.dump(newBase,file)
        except Exception as errMsg:
            return FALSE,ERR_FILE_NOT_EXIST,errMsg
        return TRUE,ERR_NO_ERRO,STRNULL

    def getJsonContent(self):
        try:    
            with open(self.pathFile,'r') as file:
                oldContent=json.load(file)
        except Exception as errMsg:
            return FALSE,ERR_FILE_NOT_EXIST,errMsg,DCTNULL
        return TRUE,ERR_NO_ERRO,STRNULL,oldContent

    def isJsonEmpty(self):
        work,errNum,errMsg,content=self.getJsonContent()
        if work==FALSE:
            return NULL
        if content==DCTNULL:
            return TRUE
        return FALSE
    
    def deleteDataBase(self):
        work,codErr,msgErr=self.deleteJson()
        if work==FALSE:
            return work,ERR_DATABASE_NOT_DELETED,msgErr        
        self.createDataBase()
        return TRUE,ERR_NO_ERRO,STRNULL

    def isTableEmpty(self,tableName):
        tblName=tableName.upper()
        if self.isJsonEmpty():
            return TRUE
        work,errNum,errMsg,base=self.getJsonContent()
        if not work:
            return TRUE

        if checkKeys(tblName,base)==FALSE:
            return TRUE

        if base[tableName]==LSTNULL:
            return TRUE

        return FALSE

    def createTable(self,tableName,tblFields,tblStdTypeValue):
        tblName=str(tableName).upper()
        if len(tblFields)!=len(tblStdTypeValue):
            return FALSE,ERR_LEN_TABLES,STRNULL        
        work,codErr,msgErr,previousContent=self.getJsonContent()
        if work==FALSE:
            return work,codErr,msgErr              
        if tblName in previousContent.keys():
            return FALSE,ERR_TABLE_ALREADY_EXIST,STRNULL
        dictTable={'ID':-1}
        position=ZERO
        for eachElement in tblFields:
            dictTable[str(eachElement).upper()]=tblStdTypeValue[position]
            position+=1           
        lstRegister=[dictTable]
        previousContent[tblName]=lstRegister
        work,codErr,msgErr=self.updateJson(previousContent)
        return work,codErr,msgErr
    
    def getAllTablesNames(self):
        work,codErr,msgErr,base=self.getJsonContent()
        if work==FALSE:
            return work,codErr,msgErr,LSTNULL        
        return work,codErr,msgErr,list(base.keys())

    def getTable(self,tableName):
        work,codErro,msgErro,base=self.getJsonContent()
        if work==FALSE:
            return work,codErro,msgErro,LSTNULL        
        tblName=tableName.upper()        
        if checkKeys(tblName,base)==FALSE:
            return FALSE,ERR_TABLE_NOT_EXIST,STRNULL,LSTNULL
        return TRUE,ERR_NO_ERRO,STRNULL,base[tblName]    
  
    def updateTable(self,tableName,tblData):
        work,codErro,msgErro,base=self.getJsonContent()
        if work==FALSE:
            return work,codErro,msgErro
        tblName=tableName.upper()  
        if checkKeys(tblName,base)==FALSE:
            return FALSE,ERR_TABLE_NOT_EXIST,STRNULL
        
        work,numErr,msgErr=self.deleteAllRecords(tableName)
        if not work:
            return work,numErr,msgErr
        work,numErr,msgErr,cont=self.insertNewsRecords(tblName,tblData)
        return work,numErr,msgErr,cont

    def updateTable2(self,tableName,tblData):
        work,codErro,msgErro,base=self.getJsonContent()
        if work==FALSE:
            return work,codErro,msgErro
        tblName=tableName.upper()  
        if checkKeys(tblName,base)==FALSE:
            return FALSE,ERR_TABLE_NOT_EXIST,STRNULL
        base[tblName]=tblData       
        
        work,codErro,msgErro=self.updateJson(base)
        if work==FALSE:
            return work,codErro,msgErro

        return self.__updateIDS__(tblName)

    def getTableFields(self,tableName):
        work,codErr,msgErr,table=self.getTable(tableName)
        if work==FALSE:
            return work,codErr,msgErr,LSTNULL        
        return work,codErr,msgErr, list(table[0].keys())        

    def subscribeTable(self,tableName,tableValues):        
        work,codErro,msgErro,base=self.getJsonContent()
        if work==FALSE:
            return work,codErro,msgErro
        tblName=tableName.upper()
        if checkKeys(tblName,base)==FALSE:
            return FALSE,ERR_TABLE_NOT_EXIST,STRNULL
        base[tblName]=tableValues
        return self.updateJson(base)
    
    def deleteTable(self,tableName):
        work,codErr,msgErr,base=self.getJsonContent()
        if work==FALSE:
            return work,codErr,msgErr
        tblName=tableName.upper()
        if checkKeys(tblName,base)==FALSE:
            return FALSE,ERR_TABLE_NOT_EXIST,STRNULL
        base.pop(tblName)
        return self.updateJson(base)     

    def deleteAllRecords(self,tableName):
        work,numErr,msgErr,lastRec=self.getLastRecord(tableName)
        if work==FALSE:
            return work,numErr,msgErr

        lastRec['ID']=-1
        work,numErr,msgErr=self.deleteTable(tableName)
        if not work:
            return work,numErr,msgErr
        fields=list(lastRec.keys())[1::]
        types=list(lastRec.values())[1::]
        return self.createTable(tableName,fields,types)
    
    def getTableLenght(self,tableName):
        work,numErr,msgErr,lastRec=self.getLastRecord(tableName)
        if work==FALSE:
            return work,numErr,msgErr,ZERO        
        return work,numErr,msgErr,lastRec['ID']+1

    def getLastRecord(self,tableName):
        tblName=tableName.upper()
        WORK,NUMERRO,MSGERRO,content=self.getJsonContent()
        if WORK==FALSE:
            return WORK,NUMERRO,MSGERRO,DCTNULL
        if checkKeys(tblName,content)==FALSE:
            return FALSE,ERR_TABLE_NOT_EXIST,STRNULL,DCTNULL
        lstLenght=len(content[tblName])
        try:
            lastRecord=content[tblName].pop(lstLenght-1)
        except Exception as MSGERRO:
            return FALSE,ERR_EMPTY_TABLE,MSGERRO,DCTNULL
        return WORK,NUMERRO,MSGERRO,lastRecord
    
    def __insertFirstReg__(self,tblName,firstRecord):
        record=[__concateDict__({'ID':0}, firstRecord)]
        return self.subscribeTable(tblName, record)

    def __validateInsertData__(self,tblName,newRecord):
        work,numErr,msgErr,lastRec=self.getLastRecord(tblName)
        if work==FALSE:
            return  work,numErr,msgErr,DCTNULL
        if len(newRecord.keys())!=len(lastRec.keys())-1:
            return FALSE,ERR_INCORRECT_DATA,'Dados não permitidos',DCTNULL        
        posicao=1
        lastRecKeys=list(lastRec.keys())
        adjustedRec=DCTNULL
        for eachElement in newRecord.keys():
            if lastRecKeys[posicao]!=str(eachElement).upper():
                return FALSE,ERR_INCORRECT_DATA,STRNULL,DCTNULL
            else:
                adjustedRec[str(eachElement).upper()]=newRecord[eachElement]
            posicao+=1            
        return TRUE,ERR_NO_ERRO,STRNULL,adjustedRec        

    def insertNewRecord(self,tableName,newRecord):
        tblName=tableName.upper()
        work,numErr,msgErr,lastRec=self.getLastRecord(tblName)
        if work==FALSE:
            return  work,numErr,msgErr        
        work,numErr,msgErr,adjustedRec=self.__validateInsertData__(tblName, newRecord)
        if work==FALSE:
            return  work,numErr,msgErr

        if lastRec['ID']==-1:
            return self.__insertFirstReg__(tblName, adjustedRec)
        
        lastId=lastRec['ID']+1
        newRecordBd=__concateDict__({'ID':lastId},adjustedRec)

        work,numErr,msgErr,table=self.getTable(tblName)
        if work==FALSE:
            return  work,numErr,msgErr
        
        table.append(newRecordBd)
        return self.subscribeTable(tblName, table)   

    def insertNewsRecords(self,tblName,lstData):
        cont=0
        if lstData==LSTNULL:
            return TRUE,ERR_NO_ERRO,STRNULL,ZERO
        for eachRecord in lstData:
            work,numErr,msgErr=self.insertNewRecord(tblName,eachRecord)
            if work==FALSE:
                return work,numErr,msgErr,cont
            cont+=1
        return TRUE,ERR_NO_ERRO,STRNULL,cont
        
    def getRecord(self,tableName,tableCriterionField,tabelCriterionValue,occurrence=1):
        work,codErr,msgErr,table=self.getTable(tableName)
        if work==FALSE:
            return work,codErr,msgErr,DCTNULL        
        
        criterionField=tableCriterionField.upper()
        criterionValue=tabelCriterionValue

        work=checkKeys(criterionField, table[0])
        if work==FALSE:
            return work,ERR_INCORRECT_FIELDS,msgErr,DCTNULL        
        turn=1
        record=DCTNULL
        for eachElement in table:
            if eachElement[criterionField]==criterionValue:
                record=eachElement
                if turn==occurrence:
                    return TRUE,ERR_NO_ERRO,STRNULL,record
                turn+=1
        
        return TRUE,ERR_NO_ERRO,STRNULL,record  

    def getRecords(self,tableName):
        work,codErr,msgErr,table=self.getTable(tableName)
        if work==FALSE:
            return work,codErr,msgErr,LSTNULL   
        tbl=[]
        for eachElement in table:
            eachElement.pop("ID")
            tbl.append(eachElement)
        return TRUE,ERR_NO_ERRO,STRNULL,tbl
    
    def checkRecord(self,tableName,tableCriterionField,tabelCriterionValue):
        work,codErr,msgErr,oldRecord=self.getRecord(tableName, tableCriterionField, tabelCriterionValue,1)
        return work

    def updateRecord(self,tableName,newData,tableCriterionField,tabelCriterionValue,occurrence=1):
        work,codErr,msgErr,oldRecord=self.getRecord(tableName, tableCriterionField, tabelCriterionValue,occurrence)
        if work==FALSE:
            return work,codErr,msgErr        
        posicao=oldRecord['ID']

        work,codErr,msgErr,table=self.getTable(tableName)
        if work==FALSE:
            return work,codErr,msgErr
        
        work,numErr,msgErr,adjustedRec=self.__validateInsertData__(tableName, newData)
        if work==FALSE:
            return work,codErr,msgErr

        table[posicao]=__concateDict__({'ID':posicao},adjustedRec)
        return self.subscribeTable(tableName, table)
    
    def __updateIDS__(self,tableName):
        work,codErr,msgErr,tabela=self.getTable(tableName)
        if work==FALSE:
            return work,codErr,msgErr        
        posicao=0
        for eachElement in tabela:
            tabela[posicao]['ID']=posicao    
            posicao+=1  
        return self.subscribeTable(tableName, tabela)

    def deleteRecord(self,tableName,tableCriterionField,tabelCriterionValue,occurrence=1):
        work,codErr,msgErr,record=self.getRecord(tableName,tableCriterionField,tabelCriterionValue,occurrence)
        if work==FALSE:
            return work,codErr,msgErr
        posicao=record['ID']
        work,codErr,msgErr,tabela=self.getTable(tableName)
        if work==FALSE:
            return work,codErr,msgErr        
        tabela.pop(posicao)
        work,codErr,msgErr=self.subscribeTable(tableName, tabela)
        if work==FALSE:
            return work,codErr,msgErr        
        return self.__updateIDS__(tableName)

class manageXls:
    def __init__(self,path=PATH_STANDARD,name="SAIDA_KLASSMATT.xlsx"):
        self.path=path + "\\" + name
        return

    def setPath(self,path,name):
        self.path=path +'\\'+ name
        return

    def create(self,title="DADOS"):
        book=xl.Workbook()
        book.worksheets[0].title=title
        book.save(self.path)
        return

    def recordXls(self,data):
        try:
            self.create()
        except Exception as errMsg:
            return FALSE,ERR_WORKBOOK_NOT_CREATED,errMsg

        book=xl.load_workbook(self.path,FALSE)
        sheet=book.worksheets[0]

        if data==LSTNULL:
            book.save(self.path)
            return TRUE,ERR_NO_ERRO,STRNULL

        column=1
        for eachElement in data[0].keys():
            sheet.cell(1,column).value=eachElement
            column+=1

        row=2
        for eachElement in data:
            column=1
            for eachOne in eachElement.values():
               sheet.cell(row,column).value=eachOne
               column+=1
            row+=1
        book.save(self.path)
        return TRUE,ERR_NO_ERRO,STRNULL

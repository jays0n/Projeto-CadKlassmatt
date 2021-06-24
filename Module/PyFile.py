import openpyxl as xl
import PyDataBaseJson as pj
import random as rd
from PyDefines import *

#standard fields:
LSTFIELDS=['NOME','DESCRICAO','MATSERV','UNIDADE','GRUPO','SUBGRUPO','URGENTE','USO','CAMINHO','CODIGO']



def sinHash(forma,valor):
    aux=['z','s','a','$','O','p',"w",'@','w',">",'<','#','%']
    if forma==1: #Oculta senha
        senHash=''        
        texto=str(valor)
        for elem in texto:
            senHash=str(aux[rd.randrange(len(aux))]) + senHash + str(hex(ord(elem)+3))+str(aux[rd.randrange(len(aux))])
        senHash=senHash.replace('x','u&')

    else: #Converte Senha
        texto=str(valor)
        for elem in aux:
            texto=str(texto.replace(elem,'')).strip()
        texto=texto.replace('u&','x')
        senHash=''
        
        for k in range(len(texto)):
            if (k+1)%4==0:
                elem=texto[(k-3):(k+1)]  
                senHash=senHash+chr(int(elem,16)-3)

    return senHash 

class readExternalFiles:
    def __init__(self,path,fileType=STRNULL,lstFields=LSTFIELDS):
        self.path=path
        self.fields=lstFields
        self.fileType=self.__getFileType__()

    def showData(self):
        data=self.getData()
        print(data)
    
    def getData(self):
        data=self.__getData__()
        if data[0]==FALSE:
            return DCTNULL
        return data[2]

    def __getData__(self):
        if self.fileType in (FILE_CSV,FILE_TXT):
            return TRUE, ERR_NO_ERRO,self.__getTxtCsvData__()
        elif self.fileType==FILE_XLS:
            return FALSE,ERR_XLS_NOT_DENIED,DCTNULL
        elif self.fileType in (FILE_XLSM,FILE_XLSX):
            return TRUE,ERR_NO_ERRO,self.__getXlsData__()
        elif self.fileType==FILE_JSON:
            return TRUE,ERR_NO_ERRO,self.__getJsonData__()
        else:
            return FALSE,ERR_FILE_TYPE_INVALID,DCTNULL    

    def __getFileType__(self):
        invertedPath=self.path[::-1]
        position=invertedPath.find('.')
        fileType=str.lower(invertedPath[:position][::-1])
        if fileType in (FILE_CSV,FILE_JSON,FILE_TXT,FILE_XLS,FILE_XLSM,FILE_XLSX):
            return fileType
        else:
            return ERR_FILE_TYPE_INVALID
    
    def __getJsonData__(self):
        fileJson=pj.manageJsonBd(self.path)
        work,codErro,msgbErro,dados=fileJson.getJsonContent()
        if work==TRUE:
            return dados
        else:
            return ERR_FILE_EMPTY
    
    def __getTxtCsvData__(self):
        dados={}
        fields=self.fields
        with open(self.path,'r') as file:
            linha=0
            for eachElement in file:       
                dctDados={}                    
                if eachElement not in (NULL,STRNULL):
                    posicao=0
                    for eachValue in eachElement.split(';'):
                        if eachValue in (NULL,STRNULL):               
                            dctDados[str(fields[posicao])]=1
                        dctDados[str(fields[posicao])]=eachValue
                        posicao+=1
                    dados['linha'+str(linha+1)]=dctDados
                    linha+=1
        return dados

    def __getXlsData__(self):
        wBook=xl.open(self.path)
        sheetName=wBook.sheetnames[0]
        wSheet=wBook[sheetName]
        linha=1
        dados={}
        fields=self.fields
        numColumns=len(fields)
        while wSheet.cell(linha, 1).value not in (NULL,STRNULL,LSTNULL):
            dctDados={}
            for coluna in range(numColumns):
                dctDados[fields[coluna]]=str(wSheet.cell(linha, coluna+1).value).replace('\n','')
            dados['linha'+str(linha)]=dctDados
            linha+=1        
        return dados
        
class importDataRecord(readExternalFiles):
    def __init__(self, path, fileType=STRNULL, lstFields=LSTFIELDS):
        super().__init__(path, fileType=fileType, lstFields=lstFields)

    def getData(self):
        return self.__formatData__()

    def __formatData__(self):
        dctData=self.__getData__()[2]
        lstData=LSTNULL
        for key in dctData.keys():
            eachElement=dctData[key]
            if str(eachElement['urgente']).upper() in ('TRUE','URGENTE','URG','VERDADE','VERDADEIRO'):
                eachElement['urgente']=TRUE
            else:
                eachElement['urgente']=FALSE

            paths=[]
            for eachPath in str(eachElement['caminho']).split(';'):
                paths.append(eachPath)
            eachElement['caminho']=paths

            eachElement['descricao']=str(str(eachElement['descricao']).upper()).replace('_X000D_',' \ ')
            lstData.append(eachElement)

        return lstData
  
class manageDataSettings(readExternalFiles):

    def getDataTable(self,tableName):
        jsonManager=pj.manageJsonBd(self.path)
        response=jsonManager.getTable(str(tableName).upper())
        if response[0]==FALSE:
            return DCTNULL
        return response[3]

    def getInitialMenuSettings(self):
        return self.getDataTable('MENUSETTINGS')[0]

    def getLoginSaved(self):
        userId=self.getInitialMenuSettings()['IDUSERRECORDED']
        if userId not in (NULL,STRNULL):
            jsonManager=pj.manageJsonBd(self.path)
            userInfo=jsonManager.getRecord('USERDATA','ID',userId)
            if userInfo[0]==FALSE:
                return DCTNULL
            return userInfo[3]
        return DCTNULL

    def __getLogin__(self):
        dctData=self.getLoginSaved()
        if dctData==DCTNULL:
            return LSTNULL
        lstData=[]
        lstData.append(dctData['USER'])
        lstData.append(sinHash(2,dctData['PASSWORD']))
        return lstData


    def getMainSettings(self):
        jsonManager=pj.manageJsonBd(self.path)
        response=jsonManager.getRecord('MAINSETTINGS','ID',0)
        if response[0]==FALSE:
            return DCTNULL
        return response[3]

    def getUrl(self):
        dctData=self.getMainSettings()
        return dctData['URL']

    def saveUser(self,user=STRNULL,password=STRNULL):
        jsonManager=pj.manageJsonBd(self.path)
        response=jsonManager.getRecord('USERDATA','USER',user)
        newRecord=DCTNULL
        newRecord['USER']=str(user).upper()
        newRecord['PASSWORD']=sinHash(1,password)
        if response[0]==FALSE:            
            response=jsonManager.insertNewRecord('USERDATA',newRecord)
        else:
            response=jsonManager.updateRecord('USERDATA',newRecord,'USER',user)
        return

    def getDataToRecord(self):
        lstData=list(self.getData().values())
        return lstData

class dataFromRecordsXls(readExternalFiles):
    def __init__(self, path, fileType=STRNULL, lstFields=LSTFIELDS):
        super().__init__(path, fileType=fileType, lstFields=lstFields)

    def __getXlsData__(self):
        wBook=xl.open(self.path)
        sheetName=wBook.sheetnames[0]
        wSheet=wBook[sheetName]
        
        initialLine=1
        while str(wSheet.cell(initialLine,1).value).upper().find('NOME')==EOF:            
            initialLine+=1   

        lastColumn=1
        while wSheet.cell(initialLine,lastColumn).value not in (NULL,STRNULL,LSTNULL):
              lastColumn+=1

        data=[]
        initialLine+=1
        lastColumn+=1
        while wSheet.cell(initialLine, 1).value not in (NULL,STRNULL,LSTNULL):
            lstData=[]
            for column in range(1,lastColumn):
                cellValue=str(wSheet.cell(initialLine,column).value).strip()
                lstData.append(cellValue)
            data.append(lstData)
            initialLine+=1        
        return data
